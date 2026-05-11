import os
import sys
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from exporter import COLUMNS

_BASE_DIR = os.path.dirname(os.path.abspath(__file__))
RESULTS_DIR = os.path.join(_BASE_DIR, "results")


def load_file(path):
    wb = load_workbook(path)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    header = list(rows[0])
    return [dict(zip(header, row)) for row in rows[1:]]


def get_cities(files):
    cities = set()
    for f in files:
        name = os.path.basename(f)
        if name.startswith("master_"):
            continue
        city = name.split("_")[0]
        if city:
            cities.add(city)
    return cities


def merge_city(city, files):
    city_files = [
        f for f in files
        if os.path.basename(f).startswith(city + "_")
        and not os.path.basename(f).startswith("master_")
    ]

    if not city_files:
        print(f"No files found for city: {city}")
        return

    all_rows = []
    for f in city_files:
        all_rows.extend(load_file(f))

    seen = set()
    unique = []
    for row in all_rows:
        key = (row.get("Company Name", "") or "", row.get("Phone", "") or "")
        if key not in seen:
            seen.add(key)
            unique.append(row)

    out_path = os.path.join(RESULTS_DIR, f"master_{city}.xlsx")
    wb = Workbook()
    ws = wb.active
    ws.title = "Master"

    ws.append(COLUMNS)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for row in unique:
        ws.append([row.get(col, "") for col in COLUMNS])

    wb.save(out_path)
    print(f"{city}: {len(unique)} unique results → {out_path}")


def main():
    if not os.path.isdir(RESULTS_DIR):
        print("No results/ folder found. Run main.py first.")
        return

    files = [
        os.path.join(RESULTS_DIR, f)
        for f in os.listdir(RESULTS_DIR)
        if f.endswith(".xlsx")
    ]

    if not files:
        print("No .xlsx files found in results/.")
        return

    if len(sys.argv) > 1:
        cities = sys.argv[1:]
    else:
        cities = sorted(get_cities(files))

    if not cities:
        print("No cities detected.")
        return

    for city in cities:
        merge_city(city, files)


if __name__ == "__main__":
    main()
